import os
import re
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ArxmlToExcelConverter:
    def __init__(self, arxml_file: str, output_excel: str):
        self.arxml_file = arxml_file
        self.output_excel = output_excel
        self.tree = None
        self.ns = {'ns': 'http://autosar.org/schema/r4.0'}
        
    def _load_arxml(self):
        """加载 ARXML 文件"""
        print(f"🔍 加载 ARXML 文件: {self.arxml_file}")
        try:
            with open(self.arxml_file, 'rb') as f:
                self.tree = etree.parse(f)
            print("✅ ARXML 文件加载成功")
        except Exception as e:
            print(f"❌ 加载 ARXML 文件失败: {e}")
            raise
    
    def _get_element_text(self, element, tag_name):
        """获取命名空间元素的文本"""
        if element is not None:
            elem = element.find(f"{{{self.ns['ns']}}}{tag_name}")
            return elem.text if elem is not None else None
        return None
    
    def _get_type_reference(self, element):
        """从 TYPE-TREF 或 IMPLEMENTATION-DATA-TYPE-REF 获取类型引用"""
        # 尝试获取 APPLICATION-PRIMITIVE-DATA-TYPE 引用
        tref = element.find(".//ns:TYPE-TREF", self.ns)
        if tref is not None and tref.text:
            return tref.text.split('/')[-1]
        
        # 尝试获取 IMPLEMENTATION-DATA-TYPE 引用
        idt_ref = element.find(".//ns:IMPLEMENTATION-DATA-TYPE-REF", self.ns)
        if idt_ref is not None and idt_ref.text:
            return idt_ref.text.split('/')[-1]
        
        return None
    
    def extract_types(self):
        """提取类型定义"""
        print("🔍 提取类型定义...")
        types_data = []
        
        # 提取 APPLICATION-PRIMITIVE-DATA-TYPE
        primitive_types = self.tree.xpath("//ns:APPLICATION-PRIMITIVE-DATA-TYPE", namespaces=self.ns)
        print(f"  - 找到 {len(primitive_types)} 个基本类型")
        for prim in primitive_types:
            name = self._get_element_text(prim, "SHORT-NAME")
            if name and name in ['uint8','Uint8','UInt8', 'uint16', 'Uint16', 'UInt16', 'uint32', 'Uint32','UInt32','uint64', 'Uint64', 'UInt64', 'sint8', 'Sint8', 'Sint8', 'sint16', 'Sint16', 'Sint16', 'sint32', 'Sint32', 'Sint32', 'sint64', 'Sint64', 'Sint64', 'float32', 'Float32', 'Float32', 'float64', 'Float64', 'Float64', 'boolean','Boolean']:
                types_data.append([name, "basic", "", name])
        # 提取 APPLICATION-ARRAY-DATA-TYPE
        array_types = self.tree.xpath("//ns:APPLICATION-ARRAY-DATA-TYPE", namespaces=self.ns)
        print(f"  - 找到 {len(array_types)} 个数组类型")
        for arr in array_types:
            name = self._get_element_text(arr, "SHORT-NAME")
            element = arr.find("ns:ELEMENT", self.ns)
            if element is not None:
                base_type = self._get_type_reference(element)
                max_elements = self._get_element_text(element, "MAX-NUMBER-OF-ELEMENTS")
                if name and base_type and max_elements:
                    array_def = f"{base_type}[{max_elements}]"
                    types_data.append([name, "array", "", array_def])
        
        # 提取 APPLICATION-RECORD-DATA-TYPE
        record_types = self.tree.xpath("//ns:APPLICATION-RECORD-DATA-TYPE", namespaces=self.ns)
        print(f"  - 找到 {len(record_types)} 个结构体类型")
        for rec in record_types:
            name = self._get_element_text(rec, "SHORT-NAME")
            elements = rec.find("ns:ELEMENTS", self.ns)
            if elements is not None:
                elements = elements.findall("ns:APPLICATION-RECORD-ELEMENT", self.ns)
                print(f"    - 结构体 {name} 有 {len(elements)} 个字段")
                for i, elem in enumerate(elements):
                    field_name = self._get_element_text(elem, "SHORT-NAME")
                    field_type = self._get_type_reference(elem)
                    if field_name and field_type:
                        if i == 0:
                            # 第一个字段，包含类型名和类别
                            types_data.append([name, "struct", field_name, field_type])
                        else:
                            # 后续字段，只添加字段名和类型
                            types_data.append(["", "", field_name, field_type])
        
        print(f"✅ 提取了 {len(types_data)} 条类型数据")
        return types_data
    
    def _extract_runnable_entities(self):
        """提取 RUNNABLE-ENTITY 信息 - 从 RUNNABLES 中获取可运行实体"""
        print("🔍 提取可运行实体信息...")
        runnable_entities = {}
        
        # 查找所有 RUNNABLE-ENTITY
        runnables = self.tree.xpath("//ns:RUNNABLE-ENTITY", namespaces=self.ns)
        print(f"  - 找到 {len(runnables)} 个可运行实体")
        
        for runnable in runnables:
            runnable_name = self._get_element_text(runnable, "SHORT-NAME")
            if not runnable_name:
                continue
            
            # 存储可运行实体的基本信息
            runnable_entities[runnable_name] = {
                'name': runnable_name,
                'ports': []
            }
            
            # 检查数据接收点
            receive_points = runnable.xpath("ns:DATA-RECEIVE-POINT-BY-ARGUMENTS", namespaces=self.ns)
            for point in receive_points:
                var_accesses = point.xpath("ns:VARIABLE-ACCESS", namespaces=self.ns)
                for var_access in var_accesses:
                    access_name = self._get_element_text(var_access, "SHORT-NAME")
                    accessed_var = var_access.find("ns:ACCESSED-VARIABLE", self.ns)
                    if accessed_var is not None:
                        autosar_var = accessed_var.find("ns:AUTOSAR-VARIABLE-IREF", self.ns)
                        if autosar_var is not None:
                            port_ref = autosar_var.find("ns:PORT-PROTOTYPE-REF", self.ns)
                            target_ref = autosar_var.find("ns:TARGET-DATA-PROTOTYPE-REF", self.ns)
                            
                            if port_ref is not None and target_ref is not None:
                                port_name = port_ref.text.split('/')[-1]
                                interface_name = target_ref.text.split('/')[-1]
                                
                                # 将端口信息存储到对应的 runnable 中
                                runnable_entities[runnable_name]['ports'].append({
                                    'port_name': port_name,
                                    'interface_name': interface_name,
                                    'direction': 'R'  # 接收
                                })
            
            # 检查数据发送点
            send_points = runnable.xpath("ns:DATA-SEND-POINT-BY-ARGUMENTS", namespaces=self.ns)
            for point in send_points:
                var_accesses = point.xpath("ns:VARIABLE-ACCESS", namespaces=self.ns)
                for var_access in var_accesses:
                    access_name = self._get_element_text(var_access, "SHORT-NAME")
                    accessed_var = var_access.find("ns:ACCESSED-VARIABLE", self.ns)
                    if accessed_var is not None:
                        autosar_var = accessed_var.find("ns:AUTOSAR-VARIABLE-IREF", self.ns)
                        if autosar_var is not None:
                            port_ref = autosar_var.find("ns:PORT-PROTOTYPE-REF", self.ns)
                            target_ref = autosar_var.find("ns:TARGET-DATA-PROTOTYPE-REF", self.ns)
                            
                            if port_ref is not None and target_ref is not None:
                                port_name = port_ref.text.split('/')[-1]
                                interface_name = target_ref.text.split('/')[-1]
                                
                                # 将端口信息存储到对应的 runnable 中
                                runnable_entities[runnable_name]['ports'].append({
                                    'port_name': port_name,
                                    'interface_name': interface_name,
                                    'direction': 'S'  # 发送
                                })
        
        print(f"  - 提取了 {sum(len(runnable['ports']) for runnable in runnable_entities.values())} 个端口访问")
        return runnable_entities
    
    def _extract_events_and_tasks(self):
        """提取事件和任务信息 - 从 SWC-INTERNAL-BEHAVIOR 中提取事件信息"""
        print("🔍 提取事件和任务信息...")
        event_runnable_mapping = {}
        
        # 查找所有 SWC-INTERNAL-BEHAVIOR
        behaviors = self.tree.xpath("//ns:RUNNABLES", namespaces=self.ns)
        
        for behavior in behaviors:
            # 查找所有事件
            events = behavior.xpath(".//ns:RUNNABLE-ENTITY | .//ns:DATA-RECEIVE-EVENT | .//ns:OPERATION-INVOKED-EVENT", namespaces=self.ns)
            for event in events:
                event_name = self._get_element_text(event, "SHORT-NAME")
                event_runnable_mapping[event_name] = {
                    'runnable': event_name  # 默认使用事件名作为可运行实体名
                }

        return event_runnable_mapping
    
    def extract_apis(self):
        """提取 API 接口定义"""
        print("🔍 提取 API 接口定义...")
        
        # 提取可运行实体
        runnable_entities = self._extract_runnable_entities()
        print(f"  - 可运行实体包含 {len(runnable_entities)} 个实体")
        
        # 提取事件和任务映射
        event_runnable_mapping = self._extract_events_and_tasks()
        
        # 提取 SENDER-RECEIVER-INTERFACE
        interfaces = self.tree.xpath("//ns:SENDER-RECEIVER-INTERFACE", namespaces=self.ns)
        print(f"  - 找到 {len(interfaces)} 个接口")
        
        # 创建接口映射
        interface_types = {}
        for iface in interfaces:
            iface_name = self._get_element_text(iface, "SHORT-NAME")
            data_elements = iface.find("ns:DATA-ELEMENTS", self.ns)
            if data_elements is not None:
                var_data = data_elements.find("ns:VARIABLE-DATA-PROTOTYPE", self.ns)
                if var_data is not None:
                    type_ref = self._get_type_reference(var_data)
                    if iface_name and type_ref:
                        interface_types[iface_name] = type_ref
        
        # 通过端口信息补充 API 详情
        p_ports = self.tree.xpath("//ns:P-PORT-PROTOTYPE", namespaces=self.ns)
        r_ports = self.tree.xpath("//ns:R-PORT-PROTOTYPE", namespaces=self.ns)
        
        print(f"  - 找到 {len(p_ports)} 个提供端口 (P-PORT)")
        print(f"  - 找到 {len(r_ports)} 个需求端口 (R-PORT)")
        
        # 为端口创建映射
        port_mapping = {}
        
        # 处理提供端口 (P-PORT)
        for port in p_ports:
            port_name = self._get_element_text(port, "SHORT-NAME")
            comp_spec = port.find(".//ns:PROVIDED-COM-SPECS", self.ns)
            if comp_spec is not None:
                data_ref = comp_spec.find(".//ns:DATA-ELEMENT-REF", self.ns)
                if data_ref is not None and data_ref.text:
                    interface_name = data_ref.text.split('/')[-1]
                    swc_name = self._get_swc_name_for_port(port)
                    port_mapping[interface_name] = {
                        'port_name': port_name,
                        'direction': 'S',  # Sender
                        'swc': swc_name if swc_name else ' '
                    }
        
        # 处理需求端口 (R-PORT)
        for port in r_ports:
            port_name = self._get_element_text(port, "SHORT-NAME")
            comp_spec = port.find(".//ns:REQUIRED-COM-SPECS", self.ns)
            if comp_spec is not None:
                data_ref = comp_spec.find(".//ns:DATA-ELEMENT-REF", self.ns)
                if data_ref is not None and data_ref.text:
                    interface_name = data_ref.text.split('/')[-1]
                    swc_name = self._get_swc_name_for_port(port)
                    port_mapping[interface_name] = {
                        'port_name': port_name,
                        'direction': 'R',  # Receiver
                        'swc': swc_name if swc_name else ' '
                    }
        
        
        # 根据可运行实体和事件任务映射生成 API 数据
        apis_data = []
        for port_name, port_info in port_mapping.items():
            interface_name = port_name
            type_ref = interface_types.get(interface_name, "UnknownType")
            direction = port_info['direction']
            swc = port_info['swc']
            task = " "
            
            apis_data.append([
                interface_name,
                type_ref,
                direction,
                swc,
                task
            ])
        # 遍历可运行实体
        for runnable_name, runnable_info in runnable_entities.items():
            # 获取任务信息
            task = " "
            swc = " "
            
            if runnable_name in event_runnable_mapping:
                task = event_runnable_mapping[runnable_name]['runnable']
            # 为每个端口创建 API 条目
            for port_info in runnable_info['ports']:
                interface_name = port_info['interface_name']
                direction = port_info['direction']
                
                # 从接口映射获取类型引用
                type_ref = interface_types.get(interface_name, "UnknownType")
                
                # 从端口映射获取 SWC（如果任务映射中没有）
                if interface_name in port_mapping:
                    if swc == " ":
                        swc = port_mapping[interface_name]['swc']

                    if direction == 'S' or direction == 'R':
                        direction = port_mapping[interface_name]['direction']
                
                apis_data.append([
                    interface_name,
                    type_ref,
                    direction,
                    swc,
                    task
                ])
        
        # 根据 interface_name 去重，保留非默认任务的条目
        deduplicated_apis = {}
        for api_entry in apis_data:
            interface_name = api_entry[0]
            task = api_entry[4]
            
            # 如果接口名不存在，或者当前条目的任务不是默认值而现有条目是默认值，则保留当前条目
            if (interface_name not in deduplicated_apis or 
                (task != " " and deduplicated_apis[interface_name][4] == " ")):
                deduplicated_apis[interface_name] = api_entry
        
        # 转换回列表格式
        apis_data = list(deduplicated_apis.values())

        print(f"✅ 提取了 {len(apis_data)} 条 API 数据（包含 Task 信息）")
        return apis_data
    
    def _get_swc_name_for_port(self, port_element):
        """通过端口元素获取所属 SWC 名称"""
        # 从端口向上查找 SWC 组件
        parent = port_element.getparent()
        while parent is not None:
            if parent.tag.endswith('APPLICATION-SW-COMPONENT-TYPE'):
                swc_name = self._get_element_text(parent, "SHORT-NAME")
                return swc_name
            parent = parent.getparent()
        return None
    
    def create_excel(self):
        """创建 Excel 文件"""
        print("📝 创建 Excel 文件...")
        wb = Workbook()
        
        # 删除默认工作表
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 创建 Types 工作表
        types_ws = wb.create_sheet("Types")
        types_ws.append(["Name", "Category", "Field Name", "Definition"])
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, 5):
            cell = types_ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加类型数据
        types_data = self.extract_types()
        for row_data in types_data:
            types_ws.append(row_data)
        
        # 调整列宽
        for col in range(1, 5):
            types_ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 创建 APIs 工作表
        apis_ws = wb.create_sheet("APIs")
        apis_ws.append(["API Name", "Type Reference", "Port Direction", "SWC Name", "Task"])
        
        # 设置表头样式
        for col in range(1, 6):
            cell = apis_ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加 API 数据
        apis_data = self.extract_apis()
        for row_data in apis_data:
            apis_ws.append(row_data)
        
        # 调整列宽
        for col in range(1, 6):
            apis_ws.column_dimensions[get_column_letter(col)].width = 20
        
        # 保存文件
        wb.save(self.output_excel)
        print(f"✅ Excel 文件创建完成: {self.output_excel}")
    
    def convert(self):
        """执行转换"""
        print("🔄 开始转换 ARXML 到 Excel...")
        
        # 检查输入文件
        if not os.path.exists(self.arxml_file):
            print(f"❌ 输入文件不存在: {self.arxml_file}")
            return
        
        try:
            self._load_arxml()
            self.create_excel()
            print("✅ 转换完成！")
        except Exception as e:
            print(f"❌ 转换过程中发生错误: {e}")
            import traceback
            traceback.print_exc()


# === 主程序入口 ===
if __name__ == '__main__':
    input_arxml = 'SDU.arxml'  # 输入的 ARXML 文件
    output_excel = 'converted_from_arxml2.xlsx'  # 输出的 Excel 文件
    
    # 检查输入文件是否存在
    if not os.path.exists(input_arxml):
        print(f"❌ 输入文件不存在: {input_arxml}")
        print("请确保 ARXML 文件存在。")
        print("当前目录文件列表:")
        for f in os.listdir('.'):
            if f.endswith('.arxml'):
                print(f"  - {f}")
        exit(1)
    
    converter = ArxmlToExcelConverter(input_arxml, output_excel)
    converter.convert()
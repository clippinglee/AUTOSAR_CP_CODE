import os
import re
import uuid
from lxml import etree
from openpyxl import load_workbook
from typing import List, Dict, Set, Tuple
from datetime import datetime


class SwcTaskArxmlGenerator:
    NS = 'http://autosar.org/schema/r4.0'

    BASIC_TYPES = {
        'boolean': 'boolean',
        'Boolean': 'boolean',
        'bool': 'boolean',
        'uint8': 'uint8',
        'Uint8': 'uint8',
        'UInt8': 'uint8',
        'uint16': 'uint16',
        'Uint16': 'uint16',
        'UInt16': 'uint16',
        'uint32': 'uint32',
        'Uint32': 'uint32',
        'UInt32': 'uint32',
        'uint64': 'uint64',
        'Uint64': 'uint64',
        'UInt64': 'uint64',
        'sint8': 'sint8',
        'Sint8': 'sint8',
        'SInt8': 'sint8',
        'sint16': 'sint16',
        'Sint16': 'sint16',
        'SInt16': 'sint16',
        'sint32': 'sint32',
        'Sint32': 'sint32',
        'SInt32': 'sint32',
        'sint64': 'sint64',
        'Sint64': 'sint64',
        'SInt64': 'sint64',
        'float32': 'float32',
        'Float32': 'float32',
        'float64': 'float64',
        'Float64': 'float64',
    }

    ARRAY_PATTERN = re.compile(r'^(\w+)\[(\d+)\]$')

    def __init__(self, excel_file: str, output_arxml: str, types_sheet: str = 'Types', apis_sheet: str = 'APIs'):
        self.excel_file = excel_file
        self.output_arxml = output_arxml
        self.types_sheet = types_sheet
        self.apis_sheet = apis_sheet
        self.type_definitions = {}
        self.api_entries = []  # (api_name, type_ref, port_dir, swc_name, task)
        self.created_idt_types = set()
        self.interface_created = set()

    def _generate_uuid(self) -> str:
        """生成标准 UUID4 字符串"""
        return str(uuid.uuid4()).replace('-', '').lower()

    def _read_type_definitions(self):
        wb = load_workbook(self.excel_file)
        ws = wb[self.types_sheet]
        
        # 读取表头（第1行）
        header = [cell.value for cell in ws[1]]  # 假设表头在第1行

        # 定义列名映射（必须与Excel表头完全一致）
        col_map = {
            'name': 'Name',
            'category': 'Category',
            'field_name': 'Field Name',
            'definition': 'Definition'
        }

        # 获取各列索引
        col_idx = {}
        for key, col_name in col_map.items():
            try:
                col_idx[key] = header.index(col_name)
            except ValueError:
                raise ValueError(f"Required column '{col_name}' not found in sheet '{self.types_sheet}'")
            
        # 检查是否有 Field Name 列（第3列）
        has_field_name_col = False
        # 安全获取单元格值
        def get_cell(key):
            idx = col_idx[key]
            return row[idx] if idx < len(row) else None
        
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            if get_cell('field_name') is not None:
                has_field_name_col = True
                break
        
        types = {}
        current_type = None
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, cat, field_name, defn = None, None, None, None            

            name = get_cell('name')
            cat = get_cell('category')
            field_name = get_cell('field_name')
            defn = get_cell('definition')
            
            if name:
                current_type = str(name).strip()
                if cat and str(cat).strip().lower() == "basic" and defn:
                    def_val = str(defn).strip()
                    # 不区分大小写检测自引用
                    if def_val.lower() == current_type.lower():
                        print(f"⚠️ 跳过自引用 basic 类型: {current_type}")
                        continue
                types[current_type] = (str(cat).strip().lower(), [])
                
                # 如果是 struct 且有 Field Name 列，则收集字段定义
                if str(cat).strip().lower() == "struct" and has_field_name_col:
                    if field_name and defn:
                        types[current_type][1].append((str(field_name).strip(), str(defn).strip()))
                elif str(cat).strip().lower() == "array":
                    # 数组类型，只处理 Definition 列
                    if defn:
                        types[current_type][1].append(str(defn).strip())
            elif current_type and field_name and defn and has_field_name_col:
                # 继续添加结构体字段
                types[current_type][1].append((str(field_name).strip(), str(defn).strip()))
            elif current_type and defn and not has_field_name_col:
                # 保持向后兼容：没有 Field Name 列时，继续使用 Definition 列
                types[current_type][1].append(str(defn).strip())
        
        # 如果没有 Field Name 列，需要重新处理 struct 类型
        if not has_field_name_col:
            for tname, (cat, defs) in types.items():
                if cat == "struct":
                    # 重新格式化为 (field_name, field_type) 元素列表
                    new_defs = []
                    for i, defn in enumerate(defs):
                        # 支持 "FieldName: FieldType" 格式
                        field_parts = defn.split(':', 1) if isinstance(defn, str) else (defn, defn)
                        if isinstance(defn, str) and len(field_parts) == 2:
                            field_name = field_parts[0].strip()
                            field_type = field_parts[1].strip()
                        else:
                            # 保持向后兼容
                            field_name = f"Field{i+1}"
                            field_type = defn
                        new_defs.append((field_name, field_type))
                    types[tname] = (cat, new_defs)
        
        self.type_definitions = types

    def _read_api_definitions(self):
        wb = load_workbook(self.excel_file)
        ws = wb[self.apis_sheet]

        # 读取表头（第1行）
        header = [cell.value for cell in ws[1]]  # 假设表头在第1行

        # 定义所需列名
        required_columns = {
            'api_name': 'API Name',
            'type_ref': 'Type Reference',
            'port_dir': 'Port Direction',
            'swc_name': 'SWC Name',
            'task': 'Task',
            'init_value': 'InitValue'
        }

        # 构建列名到索引的映射（0-based）
        col_index = {}
        for key, col_name in required_columns.items():
            try:
                col_index[key] = header.index(col_name)
            except ValueError:
                if key in ('init_value',):  # 可选字段
                    col_index[key] = None
                else:
                    raise ValueError(f"Required column '{col_name}' not found in sheet '{self.apis_sheet}'")

        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过空行或关键字段缺失的行
            api_name_cell = row[col_index['api_name']] if col_index['api_name'] < len(row) else None
            type_ref_cell = row[col_index['type_ref']] if col_index['type_ref'] < len(row) else None

            if not api_name_cell or not type_ref_cell:
                continue

            def safe_get_value(key):
                idx = col_index.get(key)
                if idx is None or idx >= len(row):
                    return None
                val = row[idx]
                return str(val).strip() if val is not None else None

            api_name = str(api_name_cell).strip()
            type_ref = str(type_ref_cell).strip()
            port_dir = (safe_get_value('port_dir') or 'S').upper()
            swc_name = safe_get_value('swc_name') or 'DefaultSWC'
            task = safe_get_value('task') or 'DefaultTask'
            init_value = row[col_index['init_value']] if col_index['init_value'] is not None and col_index['init_value'] < len(row) else None

            entries.append((api_name, type_ref, port_dir, swc_name, task, init_value))

        self.api_entries = entries

    # === IDT (Implementation) ===
    def _create_idt(self, name: str, category: str, sub_elements=None):
        e = etree.Element(f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE")
        e.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(e, f"{{{self.NS}}}SHORT-NAME").text = name
        if category in self.BASIC_TYPES:
            etree.SubElement(e, f"{{{self.NS}}}CATEGORY").text = self.BASIC_TYPES[category]
        elif category == "struct":
            etree.SubElement(e, f"{{{self.NS}}}CATEGORY").text = "STRUCTURE"
            if sub_elements:
                se = etree.SubElement(e, f"{{{self.NS}}}SUB-ELEMENTS")
                for field_name, field_type in sub_elements:
                    elem = etree.SubElement(se, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-ELEMENT")
                    elem.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
                    etree.SubElement(elem, f"{{{self.NS}}}SHORT-NAME").text = field_name
                    etree.SubElement(elem, f"{{{self.NS}}}CATEGORY").text = "TYPE_REFERENCE"  # ✅ 标准格式
                    props = etree.SubElement(elem, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
                    variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
                    cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
                    # ✅ 改为 IMPLEMENTATION-DATA-TYPE-REF
                    ref = etree.SubElement(cond, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-REF")
                    ref.set("DEST", "IMPLEMENTATION-DATA-TYPE")
                    ref.text = f"/DataTypes/ImplementationDataTypes/{field_type}"
        return e

    def _create_array_idt(self, array_name, base_type, size):
        e = etree.Element(f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE")
        e.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(e, f"{{{self.NS}}}SHORT-NAME").text = array_name
        etree.SubElement(e, f"{{{self.NS}}}CATEGORY").text = "TYPE_REFERENCE"  # ✅ 数组也应为 TYPE_REFERENCE
        sub = etree.SubElement(e, f"{{{self.NS}}}SUB-ELEMENTS")
        elem = etree.SubElement(sub, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-ELEMENT")
        elem.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(elem, f"{{{self.NS}}}SHORT-NAME").text = "Element"
        etree.SubElement(elem, f"{{{self.NS}}}CATEGORY").text = "TYPE_REFERENCE"
        etree.SubElement(elem, f"{{{self.NS}}}ARRAY-SIZE").text = str(size)
        props = etree.SubElement(elem, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        ref = etree.SubElement(cond, f"{{{self.NS}}}IMPLEMENTATION-DATA-TYPE-REF")
        ref.set("DEST", "IMPLEMENTATION-DATA-TYPE")
        ref.text = f"/DataTypes/ImplementationDataTypes/{base_type}"
        return e

    def _create_and_add_type(self, tname, pkg):
        if tname in self.created_idt_types:
            return
        if tname not in self.type_definitions:
            if tname in self.BASIC_TYPES or tname == 'boolean':
                pkg.append(self._create_idt(tname, tname))
                self.created_idt_types.add(tname)
            return
        cat, defs = self.type_definitions[tname]
        
        if cat == "basic":
            if defs and len(defs) > 0:
                base = defs[0] if isinstance(defs, list) and len(defs) > 0 else defs
                if isinstance(base, tuple):
                    base = base[1]
                self._create_and_add_type(base.strip(), pkg)
        elif cat == "array":
            if defs and len(defs) > 0:
                first_def = defs[0] if isinstance(defs[0], str) else defs[0][1] if isinstance(defs[0], tuple) else defs[0]
                match = self.ARRAY_PATTERN.match(first_def)
                if match:
                    base = match.group(1)
                    self._create_and_add_type(base, pkg)
                    pkg.append(self._create_array_idt(tname, base, int(match.group(2))))
                    self.created_idt_types.add(tname)
                else:
                    print(f"❌ 无法解析数组定义: {first_def}")
            else:
                print(f"❌ 数组类型 '{tname}' 缺少定义")
        elif cat == "struct":
            if defs:  # 检查是否有字段定义
                for field_def in defs:
                    field_type = field_def[1] if isinstance(field_def, tuple) else field_def
                    self._create_and_add_type(field_type.strip(), pkg)
                pkg.append(self._create_idt(tname, "struct", defs))
                self.created_idt_types.add(tname)
            else:
                print(f"⚠️ 结构体类型 '{tname}' 没有字段定义")

    # === ADT (Application) ===
    def _create_application_primitive_type(self, name: str):
        prim = etree.Element(f"{{{self.NS}}}APPLICATION-PRIMITIVE-DATA-TYPE")
        prim.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(prim, f"{{{self.NS}}}SHORT-NAME").text = name
        etree.SubElement(prim, f"{{{self.NS}}}CATEGORY").text = "VALUE"
        props = etree.SubElement(prim, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        etree.SubElement(cond, f"{{{self.NS}}}SW-CALIBRATION-ACCESS").text = "READ-WRITE"
        return prim

    def _create_application_array_type(self, array_name: str, base_type: str, size: int):
        arr = etree.Element(f"{{{self.NS}}}APPLICATION-ARRAY-DATA-TYPE")
        arr.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(arr, f"{{{self.NS}}}SHORT-NAME").text = array_name
        etree.SubElement(arr, f"{{{self.NS}}}CATEGORY").text = "ARRAY"
        props = etree.SubElement(arr, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        etree.SubElement(cond, f"{{{self.NS}}}SW-CALIBRATION-ACCESS").text = "READ-WRITE"
        elem = etree.SubElement(arr, f"{{{self.NS}}}ELEMENT")
        elem.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(elem, f"{{{self.NS}}}SHORT-NAME").text = array_name
        etree.SubElement(elem, f"{{{self.NS}}}CATEGORY").text = "VALUE"
        tref = etree.SubElement(elem, f"{{{self.NS}}}TYPE-TREF")
        tref.set("DEST", "APPLICATION-PRIMITIVE-DATA-TYPE")
        tref.text = f"/DataTypes/ApplicationDataTypes/{base_type}"
        etree.SubElement(elem, f"{{{self.NS}}}ARRAY-SIZE-SEMANTICS").text = "FIXED-SIZE"
        etree.SubElement(elem, f"{{{self.NS}}}MAX-NUMBER-OF-ELEMENTS").text = str(size)
        return arr

    def _create_application_record_type(self, struct_name: str, fields: List[Tuple[str, str]]):
        rec = etree.Element(f"{{{self.NS}}}APPLICATION-RECORD-DATA-TYPE")
        rec.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(rec, f"{{{self.NS}}}SHORT-NAME").text = struct_name
        etree.SubElement(rec, f"{{{self.NS}}}CATEGORY").text = "STRUCTURE"
        props = etree.SubElement(rec, f"{{{self.NS}}}SW-DATA-DEF-PROPS")
        variants = etree.SubElement(props, f"{{{self.NS}}}SW-DATA-DEF-PROPS-VARIANTS")
        cond = etree.SubElement(variants, f"{{{self.NS}}}SW-DATA-DEF-PROPS-CONDITIONAL")
        etree.SubElement(cond, f"{{{self.NS}}}SW-CALIBRATION-ACCESS").text = "READ-WRITE"
        element_spec = etree.SubElement(rec, f"{{{self.NS}}}ELEMENTS")
        for field_name, field_type in fields:
            comp = etree.SubElement(element_spec, f"{{{self.NS}}}APPLICATION-RECORD-ELEMENT")
            comp.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
            etree.SubElement(comp, f"{{{self.NS}}}SHORT-NAME").text = field_name
            tref = etree.SubElement(comp, f"{{{self.NS}}}TYPE-TREF")
            if self._is_array_type(field_type):
                tref.set("DEST", "APPLICATION-ARRAY-DATA-TYPE")
            elif field_type in self.type_definitions and self.type_definitions[field_type][0] == "struct":
                tref.set("DEST", "APPLICATION-RECORD-DATA-TYPE")
            else:
                tref.set("DEST", "APPLICATION-PRIMITIVE-DATA-TYPE")
            tref.text = f"/DataTypes/ApplicationDataTypes/{field_type}"
        return rec

    def _is_array_type(self, type_name: str) -> bool:
        if type_name in self.type_definitions:
            cat, _ = self.type_definitions[type_name]
            return cat == "array"
        return False

    # === Package Utilities ===
    def _ensure_package(self, root, pkg_name):
        ar_pkgs = root.find(f"{{{self.NS}}}AR-PACKAGES")
        if ar_pkgs is None:
            ar_pkgs = etree.SubElement(root, f"{{{self.NS}}}AR-PACKAGES")
        for pkg in ar_pkgs.findall(f"{{{self.NS}}}AR-PACKAGE"):
            if pkg.find(f"{{{self.NS}}}SHORT-NAME").text == pkg_name:
                elems = pkg.find(f"{{{self.NS}}}ELEMENTS")
                if elems is None:
                    elems = etree.SubElement(pkg, f"{{{self.NS}}}ELEMENTS")
                return elems
        pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
        etree.SubElement(pkg, f"{{{self.NS}}}SHORT-NAME").text = pkg_name
        elems = etree.SubElement(pkg, f"{{{self.NS}}}ELEMENTS")
        return elems

    def _ensure_subpackage(self, ar_pkgs, parent_name: str, child_name: str):
        parent_pkg = None
        for pkg in ar_pkgs.findall(f"{{{self.NS}}}AR-PACKAGE"):
            if pkg.find(f"{{{self.NS}}}SHORT-NAME").text == parent_name:
                parent_pkg = pkg
                break
        if parent_pkg is None:
            parent_pkg = etree.SubElement(ar_pkgs, f"{{{self.NS}}}AR-PACKAGE")
            etree.SubElement(parent_pkg, f"{{{self.NS}}}SHORT-NAME").text = parent_name

        for child in parent_pkg.findall(f"{{{self.NS}}}AR-PACKAGE"):
            if child.find(f"{{{self.NS}}}SHORT-NAME").text == child_name:
                elems = child.find(f"{{{self.NS}}}ELEMENTS")
                if elems is None:
                    elems = etree.SubElement(child, f"{{{self.NS}}}ELEMENTS")
                return elems

        new_child = etree.SubElement(parent_pkg, f"{{{self.NS}}}AR-PACKAGE")
        etree.SubElement(new_child, f"{{{self.NS}}}SHORT-NAME").text = child_name
        elems = etree.SubElement(new_child, f"{{{self.NS}}}ELEMENTS")
        return elems

    # === Interface & Port ===
    def _create_interface_adt(self, iface_name: str, type_name: str, default_value=None):
        iface = etree.Element(f"{{{self.NS}}}SENDER-RECEIVER-INTERFACE")
        iface.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(iface, f"{{{self.NS}}}SHORT-NAME").text = iface_name
        data_elems = etree.SubElement(iface, f"{{{self.NS}}}DATA-ELEMENTS")
        de = etree.SubElement(data_elems, f"{{{self.NS}}}VARIABLE-DATA-PROTOTYPE")
        de.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(de, f"{{{self.NS}}}SHORT-NAME").text = iface_name
        tref = etree.SubElement(de, f"{{{self.NS}}}TYPE-TREF")
        if self._is_array_type(type_name):
            tref.set("DEST", "APPLICATION-ARRAY-DATA-TYPE")
        elif type_name in self.type_definitions and self.type_definitions[type_name][0] == "struct":
            tref.set("DEST", "APPLICATION-RECORD-DATA-TYPE")
        else:
            tref.set("DEST", "APPLICATION-PRIMITIVE-DATA-TYPE")
        tref.text = f"/DataTypes/ApplicationDataTypes/{type_name}"

        # ✅ 新增：INIT-VALUE（如果提供了默认值）
        if default_value is not None:
            init_val = etree.SubElement(de, f"{{{self.NS}}}INIT-VALUE")
            # 根据类型生成对应的 VALUE-SPEC
            if self._is_array_type(type_name) or (type_name in self.type_definitions and self.type_definitions[type_name][0] == "struct"):
                # 复杂类型：使用 CONSTANT-REFERENCE（需先定义 CONSTANT）
                # 简化处理：暂不支持，或抛出警告
                print(f"⚠️ 复杂类型 '{type_name}' 的默认值暂不支持（仅支持基本类型）")
            else:
                # 基本类型：直接写数值
                num_val = etree.SubElement(init_val, f"{{{self.NS}}}NUMERICAL-VALUE-SPECIFICATION")
                etree.SubElement(num_val, f"{{{self.NS}}}SHORT-LABEL").text = f"{iface_name}_Init"
                etree.SubElement(num_val, f"{{{self.NS}}}VALUE").text = str(default_value)


        return iface

    def _create_port(self, name, iface_name, direction):
        tag = "R-PORT-PROTOTYPE" if direction == 'R' else "P-PORT-PROTOTYPE"
        port = etree.Element(f"{{{self.NS}}}{tag}")
        port.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(port, f"{{{self.NS}}}SHORT-NAME").text = name
        com_specs = etree.SubElement(port, f"{{{self.NS}}}{'REQUIRED' if direction == 'R' else 'PROVIDED'}-COM-SPECS")
        spec_tag = "NONQUEUED-RECEIVER-COM-SPEC" if direction == 'R' else "NONQUEUED-SENDER-COM-SPEC"
        spec = etree.SubElement(com_specs, f"{{{self.NS}}}{spec_tag}")
        spec.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        deref = etree.SubElement(spec, f"{{{self.NS}}}DATA-ELEMENT-REF")
        deref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
        deref.text = f"/Interfaces/{iface_name}/{iface_name}"
        etree.SubElement(spec, f"{{{self.NS}}}HANDLE-OUT-OF-RANGE").text = "NONE"
        etree.SubElement(spec, f"{{{self.NS}}}USES-END-TO-END-PROTECTION").text = "false"
        etree.SubElement(spec, f"{{{self.NS}}}ALIVE-TIMEOUT").text = "0"
        return port

    def _create_runnable_with_port_access(self, runnable_name: str, task_name: str, port_accesses: List[Tuple[str, str, str]]):
        """创建包含端口访问的 Runnable"""
        runnable = etree.Element(f"{{{self.NS}}}RUNNABLE-ENTITY")
        runnable.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
        etree.SubElement(runnable, f"{{{self.NS}}}SHORT-NAME").text = runnable_name
        
        # 添加最小启动间隔
        min_start_interval = etree.SubElement(runnable, f"{{{self.NS}}}MINIMUM-START-INTERVAL")
        min_start_interval.text = "0"
        
        # 添加并发调用信息
        concurrent = etree.SubElement(runnable, f"{{{self.NS}}}CAN-BE-INVOKED-CONCURRENTLY")
        concurrent.text = "false"
        
        # 添加端口访问点
        for port_name, port_direction, interface_name in port_accesses:
            if port_direction == 'R':  # 接收端口 - 数据接收点
                receive_points = etree.SubElement(runnable, f"{{{self.NS}}}DATA-RECEIVE-POINT-BY-ARGUMENTS")
                var_access = etree.SubElement(receive_points, f"{{{self.NS}}}VARIABLE-ACCESS")
                var_access.set("UUID", self._generate_uuid())
                
                # 生成变量访问名称
                access_name = f"IN_{port_name}_{interface_name}"
                etree.SubElement(var_access, f"{{{self.NS}}}SHORT-NAME").text = access_name
                
                accessed_var = etree.SubElement(var_access, f"{{{self.NS}}}ACCESSED-VARIABLE")
                autosar_var = etree.SubElement(accessed_var, f"{{{self.NS}}}AUTOSAR-VARIABLE-IREF")
                
                port_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}PORT-PROTOTYPE-REF")
                port_ref.set("DEST", "R-PORT-PROTOTYPE")
                port_ref.text = f"/Components/{task_name.split('_')[0] if '_' in task_name else task_name}/{port_name}"
                
                target_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}TARGET-DATA-PROTOTYPE-REF")
                target_ref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
                target_ref.text = f"/Interfaces/{interface_name}/{interface_name}"
                
            elif port_direction == 'S':  # 发送端口 - 数据发送点
                send_points = etree.SubElement(runnable, f"{{{self.NS}}}DATA-SEND-POINT-BY-ARGUMENTS")
                var_access = etree.SubElement(send_points, f"{{{self.NS}}}VARIABLE-ACCESS")
                var_access.set("UUID", self._generate_uuid())
                
                # 生成变量访问名称
                access_name = f"OUT_{port_name}_{interface_name}"
                etree.SubElement(var_access, f"{{{self.NS}}}SHORT-NAME").text = access_name
                
                accessed_var = etree.SubElement(var_access, f"{{{self.NS}}}ACCESSED-VARIABLE")
                autosar_var = etree.SubElement(accessed_var, f"{{{self.NS}}}AUTOSAR-VARIABLE-IREF")
                
                port_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}PORT-PROTOTYPE-REF")
                port_ref.set("DEST", "P-PORT-PROTOTYPE")
                port_ref.text = f"/Components/{task_name.split('_')[0] if '_' in task_name else task_name}/{port_name}"
                
                target_ref = etree.SubElement(autosar_var, f"{{{self.NS}}}TARGET-DATA-PROTOTYPE-REF")
                target_ref.set("DEST", "VARIABLE-DATA-PROTOTYPE")
                target_ref.text = f"/Interfaces/{interface_name}/{interface_name}"
        
        return runnable

    def _get_current_time_string(self):
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # === Main Generate ===
    def generate(self):
        self._read_type_definitions()
        self._read_api_definitions()

        nsmap = {
            None: self.NS,
            'xsd': 'http://www.w3.org/2001/XMLSchema',
            'xsi': 'http://www.w3.org/2001/XMLSchema-instance'
        }
        root = etree.Element(f"{{{self.NS}}}AUTOSAR", nsmap=nsmap)
        root.set("{http://www.w3.org/2001/XMLSchema-instance}schemaLocation",
                 f"{self.NS} AUTOSAR_4-3-0.xsd")

        comment = etree.Comment("support: Atech Co. tool version: 1.0.1 export time: " + 
                               self._get_current_time_string())
        root.addprevious(comment)

        ar_pkgs = root.find(f"{{{self.NS}}}AR-PACKAGES")
        if ar_pkgs is None:
            ar_pkgs = etree.SubElement(root, f"{{{self.NS}}}AR-PACKAGES")

        # Application Data Types
        adt_pkg = self._ensure_subpackage(ar_pkgs, "DataTypes", "ApplicationDataTypes")
        # Implementation Data Types
        idt_pkg = self._ensure_subpackage(ar_pkgs, "DataTypes", "ImplementationDataTypes")  # ✅ 更改包路径

        all_types_needed = set(entry[1] for entry in self.api_entries)
        created_adt = set()

        def create_adt_recursive(tname, visiting=None):
            if visiting is None:
                visiting = set()
            if tname in visiting:
                raise ValueError(f"检测到循环依赖或自引用类型: '{tname}' (路径: {visiting})")
            if tname in created_adt:
                return

            if tname not in self.type_definitions:
                if tname in self.BASIC_TYPES or tname == 'boolean':
                    adt_pkg.append(self._create_application_primitive_type(tname))
                    created_adt.add(tname)
                else:
                    print(f"⚠️ 警告：未定义的类型 '{tname}'，跳过 ADT 生成（请检查是否拼写错误或遗漏定义）")
                return

            visiting.add(tname)
            try:
                cat, defs = self.type_definitions[tname]
                if cat == "basic":
                    if defs and len(defs) > 0:
                        base = defs[0] if isinstance(defs, list) and len(defs) > 0 else defs
                        if isinstance(base, tuple):
                            base = base[1]  # 如果是 (field_name, field_type) 元组，取第二个元素
                        create_adt_recursive(base.strip(), visiting)
                elif cat == "array":
                    if defs and len(defs) > 0:
                        first_def = defs[0] if isinstance(defs[0], str) else defs[0][1] if isinstance(defs[0], tuple) else defs[0]
                        match = self.ARRAY_PATTERN.match(first_def)
                        if match:
                            base_type = match.group(1).strip()
                            create_adt_recursive(base_type, visiting)
                            adt_pkg.append(self._create_application_array_type(tname, base_type, int(match.group(2))))
                            created_adt.add(tname)
                        else:
                            print(f"❌ 无法解析数组定义: {first_def}")
                    else:
                        print(f"❌ 数组类型 '{tname}' 缺少定义")
                elif cat == "struct":
                    if defs:  # 检查是否有字段定义
                        field_list = []
                        for field_def in defs:
                            if isinstance(field_def, tuple):
                                field_name, field_type = field_def[0].strip(), field_def[1].strip()
                            else:
                                # 保持向后兼容
                                field_name = f"Field{defs.index(field_def)+1}"
                                field_type = field_def.strip()
                            create_adt_recursive(field_type, visiting)
                            field_list.append((field_name, field_type))
                        adt_pkg.append(self._create_application_record_type(tname, field_list))
                        created_adt.add(tname)
                    else:
                        print(f"⚠️ 结构体类型 '{tname}' 没有字段定义")
                else:
                    print(f"⚠️ 未知类型类别: {cat}，跳过 '{tname}'")
            finally:
                visiting.discard(tname)

        for tname in all_types_needed:
            create_adt_recursive(tname, set())

        # Generate IDT (for implementation layer)
        self.created_idt_types.clear()
        for tname in all_types_needed:
            if tname not in self.created_idt_types:
                self._create_and_add_type(tname, idt_pkg)

        # Interfaces (reference ADT)
        created_ifaces = set()
        for entry in self.api_entries:
            api_name, type_ref, _, _, _, init_val = entry
            if api_name not in created_ifaces:
                iface = self._create_interface_adt(api_name, type_ref, init_val)
                idt_pkg.append(iface)
                created_ifaces.add(api_name)

        # SWC Components - 按 SWC 分组
        swc_groups: Dict[str, List] = {}
        for entry in self.api_entries:
            _, _, _, swc, _,_ = entry
            swc_groups.setdefault(swc, []).append(entry)

        for swc_name, entries in swc_groups.items():
            comp_pkg = self._ensure_package(root, "Components")
            comp = etree.SubElement(comp_pkg, f"{{{self.NS}}}APPLICATION-SW-COMPONENT-TYPE")
            comp.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
            etree.SubElement(comp, f"{{{self.NS}}}SHORT-NAME").text = swc_name

            ports = etree.SubElement(comp, f"{{{self.NS}}}PORTS")
            for api_name, _, port_dir, _, _,_ in entries:
                ports.append(self._create_port(api_name, api_name, port_dir))

            internal = etree.SubElement(comp, f"{{{self.NS}}}INTERNAL-BEHAVIORS")
            ib = etree.SubElement(internal, f"{{{self.NS}}}SWC-INTERNAL-BEHAVIOR")
            ib.set("UUID", self._generate_uuid())  # ✅ 添加 UUID
            etree.SubElement(ib, f"{{{self.NS}}}SHORT-NAME").text = f"{swc_name}_InternalBehavior"

            # 按 Task 分组 API，每个 Task 创建一个 Runnable
            task_runnables = {}
            for api_name, _, port_dir, _, task, _ in entries:
                if task not in task_runnables:
                    task_runnables[task] = []
                task_runnables[task].append((api_name, port_dir))  # 添加端口方向信息

            # 为每个 Task 创建一个 Runnable（包含端口访问）
            runnables = etree.SubElement(ib, f"{{{self.NS}}}RUNNABLES")
            task_mappings = etree.SubElement(ib, f"{{{self.NS}}}TASK-EVENT")
            
            for task, apis in task_runnables.items():
                # 为 Task 创建一个 Runnable（包含端口访问）
                runnable_name = f"{task}"  # 使用任务名作为 Runnable 名称
                # 收集端口访问信息
                port_accesses = []
                for api_name, port_dir in apis:
                    port_accesses.append((api_name, port_dir, api_name))  # (port_name, port_direction, interface_name)
                
                runnables.append(self._create_runnable_with_port_access(runnable_name, task, port_accesses))
                
                # 创建 Task 映射
                task_mapping = etree.SubElement(task_mappings, f"{{{self.NS}}}TASK-EVENT")
                task_mapping.set("UUID", self._generate_uuid())
                runnable_ref = etree.SubElement(task_mapping, f"{{{self.NS}}}RUNNABLE-ENTITY-REF")
                runnable_ref.set("DEST", "RUNNABLE-ENTITY")
                runnable_ref.text = f"/Components/{swc_name}/{swc_name}_InternalBehavior/{runnable_name}"
                etree.SubElement(task_mapping, f"{{{self.NS}}}TASK-NAME").text = task

        with open(self.output_arxml, 'wb') as f:
            f.write(etree.tostring(root, pretty_print=True, xml_declaration=True, encoding='UTF-8'))
        print(f"✅ 生成完成！包含 {len(swc_groups)} 个 SWC，输出: {self.output_arxml}")


# === 主程序入口 ===
if __name__ == '__main__':
    input_excel = 'converted_from_arxml.xlsx'
    output_arxml = 'SwcOutput.arxml'

    if not os.path.exists(input_excel):
        print(f"❌ 输入文件不存在: {input_excel}")
        print("请确保 Excel 文件存在，并包含 'Types' 和 'APIs' 工作表。")
        exit(1)

    generator = SwcTaskArxmlGenerator(input_excel, output_arxml)
    generator.generate()